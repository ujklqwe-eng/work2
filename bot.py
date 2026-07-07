import asyncio
import json
import logging
import os
from datetime import datetime

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    WebAppInfo,
    ReplyKeyboardRemove,
    FSInputFile,
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)

BOT_TOKEN = "ВАШ_ТОКЕН_БОТА"          # токен от @BotFather
WEBAPP_URL = "https://ujklqwe-eng.github.io/work2/"  # ссылка на захостенный HTML

# ID администраторов — те, кто сможет скачивать сводную таблицу Excel.
# Узнать свой Telegram ID можно у бота @userinfobot (просто напишите ему /start)
ADMIN_IDS = [123456789]  # <-- замените на реальный(е) ID через запятую

EXCEL_FILE = "reports.xlsx"

COLUMNS = [
    "№", "Дата отчёта", "Время отправки", "Мастер (отправитель)", "Объект",
    "Виды работ", "Техника", "Монтажники (чел.)", "Монтажники (часы)",
    "Сварщики (чел.)", "Сварщики (часы)", "Фамилии сварщиков",
    "Проблемные вопросы", "Комментарий",
]

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


# ============================================
# РАБОТА С EXCEL
# ============================================
def init_excel():
    """Создаёт файл Excel с заголовками, если его ещё нет."""
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"
    ws.append(COLUMNS)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [5, 14, 16, 20, 22, 35, 30, 12, 12, 12, 12, 22, 28, 28]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(EXCEL_FILE)


def format_works(works: list) -> str:
    if not works:
        return "—"
    return "; ".join(
        f"{w.get('name')}: {w.get('qty')}{(' ' + w.get('unit')) if w.get('unit') else ''}"
        for w in works
    )


def format_transport(transport: list) -> str:
    if not transport:
        return "—"
    return "; ".join(
        f"{t.get('name')} — {t.get('hours')}ч"
        for t in transport
    )


def append_report_to_excel(fio: str, report_date: str, username: str, objects: list) -> int:
    """Добавляет по одной строке в Excel на каждый объект отчёта. Возвращает следующий свободный № для первой строки."""
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Детализация"]

    next_id = ws.max_row  # заголовок = строка 1, значит текущий max_row = следующий номер
    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    for obj in objects:
        row = [
            next_id,
            report_date,
            now_str,
            fio,
            obj.get("object", "—"),
            format_works(obj.get("works", [])),
            format_transport(obj.get("transport", [])),
            obj.get("mounters", {}).get("count") or "—",
            obj.get("mounters", {}).get("hours") or "—",
            obj.get("welders", {}).get("count") or "—",
            obj.get("welders", {}).get("hours") or "—",
            obj.get("welders", {}).get("surnames") or "—",
            obj.get("problems") or "—",
            obj.get("comment") or "—",
        ]
        ws.append(row)
        next_id += 1

    wb.save(EXCEL_FILE)
    return next_id


def format_report_text(fio: str, report_date: str, objects: list) -> str:
    lines = [
        "✅ <b>Отчёт принят</b>",
        f"👤 Мастер: {fio}",
        f"🕐 Дата: {report_date}",
        f"🏗 Объектов в отчёте: {len(objects)}",
        "",
    ]

    for i, obj in enumerate(objects, start=1):
        lines.append(f"<b>Объект #{i}: {obj.get('object', '—')}</b>")

        works = obj.get("works", [])
        if works:
            lines.append("🔧 Виды работ:")
            for w in works:
                unit = f" {w.get('unit')}" if w.get("unit") else ""
                lines.append(f"  • {w.get('name')}: {w.get('qty')}{unit}")

        transport = obj.get("transport", [])
        if transport:
            lines.append("🚗 Техника:")
            for t in transport:
                lines.append(f"  • {t.get('name')} — {t.get('hours')} ч.")

        mounters = obj.get("mounters", {})
        welders = obj.get("welders", {})
        if mounters.get("count"):
            lines.append(f"🔧 Монтажники: {mounters.get('count')} чел. — {mounters.get('hours')} ч.")
        if welders.get("count"):
            lines.append(f"🔥 Сварщики: {welders.get('count')} чел. — {welders.get('hours')} ч.")
            if welders.get("surnames"):
                lines.append(f"   Фамилии: {welders.get('surnames')}")

        if obj.get("problems"):
            lines.append(f"⚠️ Проблемы: {obj.get('problems')}")
        if obj.get("comment"):
            lines.append(f"💬 Комментарий: {obj.get('comment')}")

        lines.append("")

    return "\n".join(lines)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


# ============================================
# КОМАНДЫ
# ============================================
@dp.message(Command("start"))
async def start(message: Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[
            KeyboardButton(
                text="📋 Заполнить отчёт",
                web_app=WebAppInfo(url=WEBAPP_URL)
            )
        ]],
        resize_keyboard=True
    )
    await message.answer(
        "Привет! Нажмите кнопку ниже, чтобы заполнить отчёт по технике и персоналу.",
        reply_markup=keyboard
    )


# Этот хендлер срабатывает, когда мини-приложение вызывает Telegram.WebApp.sendData(...)
@dp.message(F.web_app_data)
async def handle_webapp_data(message: Message):
    try:
        data = json.loads(message.web_app_data.data)
    except json.JSONDecodeError:
        await message.answer("⚠️ Не удалось разобрать данные отчёта.")
        return

    fio = data.get("fio", "—")
    report_date = data.get("date", "—")
    objects = data.get("objects", [])
    username = message.from_user.username or message.from_user.full_name

    if not objects:
        await message.answer("⚠️ В отчёте нет ни одного объекта.")
        return

    append_report_to_excel(fio, report_date, username, objects)

    await message.answer(
        format_report_text(fio, report_date, objects),
        reply_markup=ReplyKeyboardRemove()
    )


# ------ АДМИНСКИЕ КОМАНДЫ ------

@dp.message(Command("excel"))
async def send_excel(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ Эта команда доступна только администраторам.")
        return

    if not os.path.exists(EXCEL_FILE):
        await message.answer("Пока нет ни одного отчёта.")
        return

    await message.answer_document(
        FSInputFile(EXCEL_FILE),
        caption="📊 Таблица со всеми отчётами (лист «Детализация»)"
    )


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
